import os 
# Libraries for export to database and excel
from openpyxl import load_workbook
import xlsxwriter

def export_to_excel(dict_to_xlsx:dict, WB_name:str, sheet_name:str):
    ''' 
    Function to export dictionary containing article details to Excel workbook.
    Each function call will append data to workbook.

    | Input: dictionary for export, Workbook name, Sheet name
    '''

    if WB_name[-5:] != '.xlsx': # add extension to workbook name
        WB_name = WB_name + '.xlsx'

    headers = ['URL','Title','Publish_Date','Content'] # Column headers
    
    if not os.path.isfile(WB_name): # create workbook if it does not exist
        book = xlsxwriter.Workbook(WB_name)
        sheet = book.add_worksheet(sheet_name)
        for (idx, header) in enumerate(headers):
            sheet.write(0, idx, header)
        book.close()
    
    loaded_book = load_workbook(WB_name)
    loaded_sheet = loaded_book[sheet_name]

    values = [dict_to_xlsx[key] for key in headers] # extract values for each key from dictionary 
    
    loaded_sheet.append(values) # write to excel file
    loaded_book.save(filename=WB_name)
